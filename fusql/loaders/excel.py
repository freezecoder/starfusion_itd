"""Excel loader for fusql test mode.

Writes fusion data to .xlsx files with formatting.
"""

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from fusql.loaders.base import BaseLoader


class ExcelLoader(BaseLoader):
    """Excel loader for test mode.
    
    Writes rows to formatted .xlsx files.
    """
    
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    
    def __init__(self, output_dir: Path):
        """Initialize Excel loader.
        
        Args:
            output_dir: Directory to write Excel files to.
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self._written_files: Dict[str, Path] = {}

    def write(self, table_name: str, rows: List[Dict[str, Any]]) -> Path:
        """Write rows to Excel file.
        
        Args:
            table_name: Name of the table (becomes filename).
            rows: List of row dictionaries.
            
        Returns:
            Path to the written Excel file.
        """
        if not rows:
            return self.output_dir / f"{table_name}.xlsx"
        
        output_path = self.output_dir / f"{table_name}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = table_name[:31]  # Excel sheet name max 31 chars
        
        # Get all columns from first row
        columns = list(rows[0].keys())
        
        # Write header row with formatting
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        
        # Write data rows
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                value = row_data.get(col_name)
                
                # Convert datetime to Excel-compatible format
                if isinstance(value, datetime):
                    value = value.isoformat()
                
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for col_idx, col_name in enumerate(columns, start=1):
            max_length = len(str(col_name))
            for row_idx in range(2, len(rows) + 2):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        
        wb.save(output_path)
        self._written_files[table_name] = output_path
        return output_path

    def table_exists(self, table: str) -> bool:
        """Check if Excel file exists for table.
        
        Args:
            table: Table name.
            
        Returns:
            True if Excel file exists, False otherwise.
        """
        return (self.output_dir / f"{table}.xlsx").exists()

    def create_table(self, table: str, schema: Dict[str, str]) -> None:
        """Create empty Excel file with headers.
        
        Args:
            table: Table name.
            schema: Dictionary mapping column names to types (ignored).
        """
        output_path = self.output_dir / f"{table}.xlsx"
        
        if not output_path.exists():
            wb = Workbook()
            ws = wb.active
            ws.title = table[:31]
            wb.save(output_path)

    def get_written_files(self) -> Dict[str, Path]:
        """Get mapping of table names to their written file paths.
        
        Returns:
            Dictionary mapping table names to file paths.
        """
        return dict(self._written_files)


def write_excel(output_path: Path, rows: List[Dict[str, Any]], sheet_name: str = "Sheet1") -> Path:
    """Write rows to Excel file at specified path.
    
    Args:
        output_path: Path for the output file (.xlsx)
        rows: List of row dictionaries
        sheet_name: Name for the worksheet
        
    Returns:
        Path to the written Excel file.
    """
    output_path = Path(output_path)
    
    if not rows:
        # Create empty workbook
        wb = Workbook()
        wb.save(output_path)
        return output_path
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    
    columns = list(rows[0].keys())
    
    # Write header
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Write data
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            value = row_data.get(col_name)
            if isinstance(value, datetime):
                value = value.isoformat()
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-adjust columns
    for col_idx in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    
    wb.save(output_path)
    return output_path
